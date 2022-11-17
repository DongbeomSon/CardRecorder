import cv2 #openCV
import imutils as imutils
import numpy as np
import pytesseract #OCR
import re #정규표현식
import requests #request html
from bs4 import BeautifulSoup #html parsing
from imutils.contours import sort_contours #edgeDetection
from imutils.perspective import four_point_transform
import pandas as pd #excel
import openpyxl as xl
import os

##----TESSERACT_OPTION------###

f = open('./config.cfg', 'r', encoding='utf-8')
t_path = f.readline()
f.close()
#pytesseract.pytesseract.tesseract_cmd = R'C:\Program Files\Tesseract-OCR\tesseract'
pytesseract.pytesseract.tesseract_cmd = t_path
config = ('-l eng --oem 3 --psm 7')


def loc(event, x, y, flags, param):
    if (event == 1):
        print(x / w)
        print(y / h)

def cap_to_digit(str):
    str = str.replace('Y', '9')
    str = str.replace('S', '9') #or 5
    str = str.replace('A', '4')
    str = str.replace('G', '6') #6 or 3
    str = str.replace('I', '1')
    str = str.replace('O', '0')

    return str

def find_pack(str):
    numpack_list = ['SR','RC','DP','EP', 'AC', 'CP']
    rtn = str
    for x in numpack_list:
        if(str[:2] == x):
            rtn = x + cap_to_digit(str[2:])
            #print("converted :: " + rtn)
            break

    return rtn


def adjust_gamma(img, g):
    img = img.astype(np.float64)
    img = ((img / 255) ** (1 / g)) * 255
    img = img.astype(np.uint8)

    return img

def adjust_contrast(img, a):
    rtn = np.clip((1 + a) * img - 128 * a, 0,255)
    rtn = rtn.astype(np.uint8)
    return rtn

def grayscale(img):
    return cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)

def findRec(org_img):
    img = imutils.resize(org_img, 500)
    ratio = org_img.shape[1] / float(img.shape[1])

    blurred = cv2.GaussianBlur(img, (5, 5), 0)
    canny = cv2.Canny(blurred, 100, 255, L2gradient=True)

    cnts = cv2.findContours(canny.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cnts = imutils.grab_contours(cnts)
    cnts = sorted(cnts, key=cv2.contourArea, reverse=True)

    findCnt = []
    # 정렬된 contours를 반복문으로 수행하며 4개의 꼭지점을 갖는 도형을 검출
    for c in cnts:
        peri = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, 0.02 * peri, True)

        # 사각형만 판별해서 저장
        if len(approx) == 4:
            findCnt.append(approx)

    # 만약 추출한 윤곽이 없을 경우 오류
    if len(findCnt) == 0:
        print("카드 인식 실패")

    cards = []
    for c in findCnt:
        temp = four_point_transform(org_img, c.reshape(4, 2) * ratio)
        if temp.shape[0] < temp.shape[1]:
            temp = cv2.rotate(temp, cv2.ROTATE_90_CLOCKWISE)
        rt = temp.shape[0] / temp.shape[1]
        if 1.30 < rt and rt < 1.55 and temp.shape[0] > 100:
            #print(rt)
            cards.append(temp)
            cards.append(cv2.rotate(temp,cv2.ROTATE_180))
        #else:
            #print("not relative" + str(rt))

    return cards
path = "./imgs"
file_list = os.listdir(path)
file_list_imgs = [file for file in file_list if file.endswith(".jpg")]
for img in file_list_imgs:
    org_img = cv2.imread('./imgs/'+img)
    #cv2.imshow("org",org_img)
    cards = findRec(org_img)
    print("total # of cards : " + str(len(cards)/2))
    t_list = []
    for idx, card in enumerate(cards):
        #cv2.imshow("card" + str(idx), card)

        h,w = card.shape[:2]
        gray = grayscale(card)

        roi = card[int(h*0.71):int(h*0.78), int(w*0.61):int(w*0.93)]
        mult = int(480/roi.shape[0])
        roi = cv2.resize(roi, dsize=(0, 0), fx=mult, fy=mult, interpolation=cv2.INTER_LINEAR)
        roi_img = roi.copy()
        roi = grayscale(roi)

        gray = roi.copy()

        #roi = cv2.adaptiveThreshold(roi, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11,0)
        roi = cv2.GaussianBlur(roi, (5, 5), 0)
        roi = cv2.threshold(roi,0,255,cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]

        (H, W) = gray.shape

        rectKernel = cv2.getStructuringElement(cv2.MORPH_RECT, (20, 10))
        sqKernel = cv2.getStructuringElement(cv2.MORPH_RECT, (50, 21))

        gray = cv2.GaussianBlur(gray, (11, 11), 0)
        blackhat = cv2.morphologyEx(gray, cv2.MORPH_BLACKHAT, rectKernel)

        grad = cv2.Sobel(blackhat, ddepth=cv2.CV_32F, dx=1, dy=0, ksize=-1)
        grad = np.absolute(grad)
        (minVal, maxVal) = (np.min(grad), np.max(grad))
        grad = (grad - minVal) / (maxVal - minVal)
        grad = (grad * 255).astype("uint8")

        grad = cv2.morphologyEx(grad, cv2.MORPH_CLOSE, rectKernel, iterations=5)
        thresh = cv2.threshold(grad, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]

        close_thresh = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, sqKernel)
        close_thresh = cv2.erode(close_thresh, None, iterations=3)
        #cv2.imshow("close"+str(idx),close_thresh)
        cnts = cv2.findContours(close_thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cnts = imutils.grab_contours(cnts)

        cnt_out = []
        if(len(cnts) == 0):
            print("영역 탐지 실패")
            continue
        cnts = sort_contours(cnts, method="top-to-bottom")[0]
        for k in cnts:
            (x, y, w, h) = cv2.boundingRect(k)
            r = w/h
            if w > 600 and r > 6:
                cnt_out = k
                break

        if(len(cnt_out) == 0):
            #print("인식실패")
            continue

        roi = cv2.drawContours(gray, [cnt_out], -1, (0,0,255),1)
        #cv2.imshow("cont"+str(idx),roi)

        margin_h = 30
        margin_w = 50
        (x, y, w, h) = cv2.boundingRect(cnt_out)


        y1 = y-margin_h if (y-margin_h > 0) else 0
        y2 = y+h+margin_h if (y+h+margin_h < gray.shape[0]) else gray.shape[0]
        x1 = x-margin_w if (x-margin_w > 0) else 0
        x2 = x+w+margin_w if (x+w+margin_w < gray.shape[1]) else gray.shape[1]

        roi = roi_img[y1:y2, x1:x2]
        #cv2.imshow("roi" + str(idx),roi)

        kernel = np.array([[-1, -1,-1],
                           [-1, 10.5, -1],
                           [-1, -1, -1]])
        #roi = cv2.filter2D(roi, -1, kernel)

        #ret, blackhat = cv2.threshold(roi,70,255, cv2.THRESH_BINARY)
        roi = cv2.medianBlur(roi, 1)
        roi = cv2.bilateralFilter(roi, 5, 100, 100)
        #roi = cv2.filter2D(roi, -1, kernel)
        text_out = []
        for i in range(0,5):
            cont = adjust_contrast(roi,0.2*(i))
            #cont = cv2.erode(cont,-np.ones((3,3)),iterations=2)
            gray = grayscale(cont)
            for j in range(41,101,12):
                binary = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, j, 0)
                binary = cv2.dilate(binary, -np.ones((3,3)), iterations=2)
                #binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
                #cv2.imshow("test"+str(i)+str(j)+"idx"+str(idx),binary)
                text = pytesseract.image_to_string(binary)
                if(text != ""):
                    #print(text + ": contrast : " + str(0.1*i))
                    text_out.append(text)

        #
        #roi = cv2.GaussianBlur(roi, (5, 5), 0)
        ##roi = cv2.medianBlur(roi, 5)
        #roi = cv2.erode(roi, (3,3), iterations=2)
        #roi = adjust_gamma(roi,3)

        #roi = thresh = cv2.threshold(roi, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]

        #cv2.imshow("real" + str(idx),roi)

        if(len(text_out) == 0):
            print("text 인식 실패")
            continue
        text_candid = []
        for text in text_out:
            text = text.replace(" ", "")
            text = text.replace("\n", "")
            text_list = text.split('KR')
            #0 O, l 1, 등 OCR이 구분하지 못하는 경우 치환하며 진행 필요
            if(len(text_list) >= 2):
                text_list[0] = re.sub(r"[^\uAC00-\uD7A30-9A-Z\s]", "", text_list[0])
                if(len(text_list[0]) == 4):
                    text_list[0] = find_pack(text_list[0])
                else:
                    text_list[0] = ""
                text_list[1] = cap_to_digit(text_list[1])
                text_list[1] = re.sub(r"[^\uAC00-\uD7A30-9\s]", "", text_list[1])
                if(len(text_list[1]) != 3):
                    text_list[1] = ""
                if(text_list[1] != "" and text_list[0] != ""):
                    text = text_list[0] + '-' + "KR" + text_list[1]
                    text_candid.append(text)

        print(text_candid)
        if(len(text_candid) != 0):
            t_list.append(max(text_candid, key=text_candid.count))
        #cardKeyword = text

    print(t_list)

    cards = []
    for idx, cardKeyword in enumerate(t_list):
        yugioh_db = "https://www.db.yugioh-card.com/yugiohdb/card_search.action?ope=1&sess=1&rp=20&" +\
                    "keyword=" + cardKeyword +\
                    "&stype=4&ctype=&othercon=2&starfr=&starto=&pscalefr=&pscaleto=&linkmarkerfr=&linkmarkerto=&link_m=2&atkfr=&atkto=&deffr=&defto=&request_locale=ko"

        req = requests.get(yugioh_db)
        html = req.text
        status = req.status_code
        is_ok = req.ok

        f = open('cInfo.html','w', encoding = 'utf-8')
        f.write(html)
        f.close()

        soup = BeautifulSoup(html, 'html.parser')

        cardInfo = soup.select(
            '#card_list > div > input.cnm'
            )
        ## my_titles는 list 객체
        Name = ""
        for t in cardInfo:
            ## Tag의 속성을 가져오기
            Name = t.get('value')

        cardIdLink = soup.select(
            '#card_list > div > input.link_value'
        )
        cardId = ""
        for t in cardIdLink:
            ## Tag의 속성을 가져오기
            print(t.text)
            cardId = t.get('value')
        if(Name == ""):
            print("검색 실패!")
        else:
            print("검색 됨 : " + Name)
            cards.append(Name)

    path = './export_sample.xlsx'
    if not os.path.exists(path):
        df = pd.DataFrame([[1]],
                              index=[cards], columns=['STOCK'])
        df.to_excel(path, sheet_name='data')
        print("excle 생성됨")
    else:

        wb = xl.load_workbook(path)
        ws = wb.active
        print("excel  수정됨")
        MAXROOP = 20000
        while(len(cards) != 0):
            idx = 1
            while(idx < MAXROOP):
                idx = idx + 1
                cd = ws["A" + str(idx)].value
                if(cd == None):
                    for i, x in enumerate(cards):

                        cnt = cards.count(x)
                        for j in range(0,cnt):
                            cards.remove(x)
                        ws.cell(row = i+idx, column=1).value = x
                        ws.cell(row = i+idx, column=2).value = cnt
                    break
                cnt = cards.count(cd)
                if(cnt > 0):
                    for i in range(0, cnt):
                        cards.remove(cd)
                    num = int(ws["B" + str(idx)].value)
                    num = cnt + num
                    ws.cell(row = idx, column=2).value = num
        wb.save(path)

#cv2.waitKey(0)
cv2.destroyAllWindows()